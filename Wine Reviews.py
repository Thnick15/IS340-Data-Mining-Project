import csv
import json
from collections import OrderedDict
from statistics import *
from openpyxl import *
from textblob import TextBlob
import matplotlib.pyplot as plt
import numpy as np
from collections import defaultdict
import operator
#LISTS
countList=[]
descList=[]
pointList=[]
priceList=[]
sentList=[]
#Dictionaries with defaultdict allows us to have multiple values for duplicate keys
country=defaultdict(list)
province=defaultdict(list)
description={}
nameDict={}
#Open Excel
wineBook = load_workbook("wineData.xlsx")
sheet1 = wineBook.active
#Reads of Columns
wineCountry = sheet1["b1"].value
wineDescript = sheet1["c1"].value
winePoint = sheet1["e1"].value
winePrice = sheet1["f1"].value
wineProvince = sheet1["g1"].value
wineName = sheet1["l1"].value
#Adds Excel Data to List
for i in range (2,120976):
                
        wineCountry = sheet1.cell(column = 2, row = i).value
        wineDescript = sheet1.cell(column = 3, row = i).value
        winePoint = sheet1.cell(column = 5, row = i).value
        winePrice = sheet1.cell(column = 6, row = i).value
        wineProvince = sheet1.cell(column = 7, row = i).value
        wineName = sheet1.cell(column = 12, row =i).value
        province[wineProvince].append(winePoint)
        country[wineCountry].append(winePoint)
        descList.append(wineDescript)
        countList.append(wineCountry)
        pointList.append(winePoint)
        priceList.append(winePrice)
        #Collects all 100 point wines in the nameDict
        if winePoint == 100:
                nameDict[wineName]=winePrice     

#header message
print("Dataset name: Wine Reviews")
print()

#Calculates Sentiment Analysis for each review and add to a list
for i in descList:
        review=TextBlob(i)
        sentList.append(review.sentiment.polarity)
        description[i]=review.sentiment.polarity
        

###PRICE FUNCTIONS 
#function that determines highest price
def maxPrice(priceList):
        maxIndex = 0
        maxValue = float("-inf")
        for i in range (len(priceList)):
                if priceList[i]>maxValue:
                         maxIndex = i
                         maxValue = priceList[i]
        return maxValue
#function that determines the lowest price
def minPrice(priceList):
        minIndex = 0
        minValue = float("inf")
        for i in range(len(priceList)):
                if priceList[i]<minValue:
                         minIndex = i
                         minValue = priceList[i]
        return minValue
#function that determines the average price
def avgPrice(priceList):
        avg = mean(priceList)
        return avg
#function that determines standard deviation of prices
def sd(priceList):
        sd = stdev(priceList)
        return sd
#function that determines mode price
def modPrice(priceList):
        try:
             mod = mode(priceList)
             return mod
        except StatisticsError: #except if there is no mode
             return None
#function that determines median price
def mediPrice(priceList):
        medi = median(priceList)
        return medi
#function that determines the range of prices
def rangeCalc(priceList):
        return max(priceList) - min(priceList)

###POINT FUNCTIONS
#function that determines highest points
def maxPoint(pointList):
        maxIndex = 0
        maxValue = float("-inf")
        for i in range (len(pointList)):
                if pointList[i]>maxValue:
                         maxIndex = i
                         maxValue = pointList[i]
        return maxValue
#function that determines the lowest points
def minPoint(pointList):
        minIndex = 0
        minValue = float("inf")
        for i in range(len(pointList)):
                if pointList[i]<minValue:
                         minIndex = i
                         minValue = pointList[i]
        return minValue
#function that determines average points
def averagePoint(pointList):
        total = 0
        for i in pointList:
                total = total + i
                avg = total/len(pointList)
        return round(avg,2)
#funcion that determines the mode of points
def modPoint(pointList):
        try:
                mod = mode(pointList)
                return mod
        except StatisticsError: #except if there is no mode
                return None
#function that determines the median of points
def mediPoint(pointList):
        medi = median(pointList)
        return medi
#function that determines the standard deviation of points
def sd(pointList):
        sd = stdev(pointList)
        return sd
#function that determines the range of points
def rangeCalc(pointList):
        return max(pointList) - min(pointList)

###SENTIMENT FUNCTIONS
#function that determines the max sentiment value
def maxSent(sentList):
        maxSent = max(sentList)
        return round(maxSent,2)
#function that determines the min sentiment value
def minSent(sentList):
        minSent = min(sentList)
        return round(minSent,2)
#function that determines the standard deviation of sent list
def sdevFunc(sentList):
        total=0
        sumDiff=0
        n=len(sentList)
        #for loop calculates the mean of the sentList
        for i in sentList:
                total=total+i
                avg=(total/n)
        #for loop calculates the sum of the squared differences
        for x in sentList:
                sumX=(x-avg)**2
                sumDiff=sumX+sumDiff
        #Calculates the standard deviation of the sentList
        sde=(((sumDiff))/(n-1))**(1/2.0)
        return round(sde,2)
#function that determines the avg of description in sentlist
def averageDesc(sentList):
        return sum(sentList)/len(sentList)
#function that determines the range of sentiment
def rangeCalc(sentList):
        return max(sentList) - min(sentList)
#avgDict to get average points for each country
avgDict = {}
print("Below is the average points for each country: ")
print()
for wineCountry,winePoint in country.items():
        avgDict[wineCountry]= round(sum(winePoint)/float(len(winePoint)),2)
for wineCountry,winePoint in avgDict.items():
        print(wineCountry, "-", winePoint)

#avgProv to get average points for each Province
avgProv = {}
print("Below is the average points for each province: ")
print()
for wineProvince,winePoint in province.items():
        avgProv[wineProvince]= round(sum(winePoint)/float(len(winePoint)),2)
for wineProvince,winePoint in avgProv.items():
        print(wineProvince, "-", winePoint)

#numerical - structured data
print()
print("--- Points Data ---")
print("Highest Points:", maxPoint(pointList))
print("Lowest Points:", minPoint(pointList))
print("Average Points: %.2f" % averagePoint(pointList))
print("Mode of Points:", modPoint(pointList))
print("Median of Points: %.2f" % mediPoint(pointList))
print("Range of Points:", rangeCalc(pointList))
print("Standard Deviation Value of Points: %.2f" % sd(pointList))
#numerical - structured data
print()
print("--- Price Data ---")
print("Highest Price: %.2f" % maxPrice(priceList))
print("Lowest Price: %.2f" % minPrice(priceList))
print("Average Price: %.2f " % avgPrice(priceList))
print("Mode of Prices:", modPrice(priceList))
print("Median of Prices:", mediPrice(priceList))
print("Range of Prices: %.2f" % rangeCalc(priceList))
print("Standard Deviation Value of Prices: %.2f" % sd(priceList))
#textual - unstructured data
print()
print("--- Description Data ---")
print("Maximum Sentimental Value of the Wine Description: %.2f" % maxSent(sentList))
print("Minimum Sentimental Value of the Wine Description: %.2f" % minSent(sentList))
print("Average Sentiment of the List : %.2f" % averageDesc(sentList))
print("Range of Sentiment: %.2f" % rangeCalc(sentList))
print("Standard Deviation Value of the Wine Description: %.2f" % sdevFunc(sentList))

#Highest Rated Country
maxCount=(max(avgDict.items(), key=operator.itemgetter(1)))
#Lowest Rated Country
minCount=(min(avgDict.items(), key=operator.itemgetter(1)))
#Highest Rated Province
maxProv=(max(avgProv.items(), key=operator.itemgetter(1)))
#Lowest Rated Province
minProv=(min(avgProv.items(), key=operator.itemgetter(1)))
#Highest Sentiment Description
maxDescript=(max(description.items(), key=operator.itemgetter(1)))
#lowest Sentiment Description
minDescript=(min(description.items(), key=operator.itemgetter(1)))
#lowest Priced 100 Point Wine
cheapWine =(min(nameDict.items(),key=operator.itemgetter(1)))

#Creates a new text file and names it f in python
with open('mohTXT.txt', 'w',encoding='utf-8') as f:
        #Writes the average country rating in the new text file
        f.write("Here is the list of all the Countries and their Average wine point rating\n")
        f.writelines('{0}:{1}\n'.format(wineCountry,winePoint) for wineCountry, winePoint in(avgDict.items()))
        f.write('\n')
        #writes the average province rating in the new text file
        f.write("Here is the list of all the Provinces and their Average wine point rating\n")
        f.writelines('{0}:{1}\n'.format(wineProvince,winePoint) for wineProvince, winePoint in(avgProv.items()))
        f.write('\n')
        #Highest Rated Country
        f.write("The Highest Rated Country is : " + str(maxCount)+'\n')
        #Lowest Rated Country
        f.write("The Lowest Rated Country is : " +str(minCount)+'\n')
        #Highest Rated Province
        f.write("The Highest Rated Province is : " +str(maxProv)+'\n')
        #Lowest Rated Province
        f.write("The Lowest Rated Province is : "+str(minProv)+'\n')
        #Highest Sentiment Description
        f.write("The Highest Sentiment is : " +str(maxDescript)+'\n')
        #lowest Sentiment Description
        f.write("The Lowest Sentiment is : " +str(minDescript)+'\n')
        #Lowest Priced 100 point wine
        f.write('\n')
        f.write("The Lowest Priced 100 Point Rated Wine is: " + str(cheapWine))
        
f.close()

#Correlation between description and the points
#Correlation between points and price
#Correlation between price and rating
#list1=priceList
#list2=pointList
#list3=sentList
#plot1=plt.scatter(list1,list2)
#plot2=plt.scatter(list1,list3)
#plot3=plt.hist(list2,list3)
#plot3=plt.scatter(list2,list3)
#plt.show(plot1)
#plt.show(plot2)
#plt.show(plot3)
